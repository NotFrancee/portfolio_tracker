import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from classes.trade import Trade
from classes.position import Position


class ExcelInterface:
    """handles data in the excel spreadsheet"""

    def __init__(self, path_to_spreadsheet: str) -> None:
        self.path = path_to_spreadsheet
        self.wb = openpyxl.load_workbook(path_to_spreadsheet)

    def read_worksheet(
        self, ws: Worksheet
    ) -> tuple[list[str], list[list[str]]]:
        """Given a worksheeets, reads and returns
        a tuple: (header, data)

        Args:
            ws (openpyxl.Worksheet)

        Returns:
            (header, rows_data): first element is the header row,
            the second element is the actual data
        """

        rows = []
        for row in ws.iter_rows():
            row_values = [cell.value for cell in row]
            rows.append(row_values)

        header: list[str] = [x.lower().replace(" ", "_") for x in rows[0]]
        data: list[list[str]] = rows[1:]

        return header, data

    def retrieve_trades(self) -> list[Trade]:
        """Retrieves Trades from the workbook,
        and returns a list of trades

        Returns:
            list[Trade]
        """

        ws = self.wb["Trades"]

        header, rows_data = self.read_worksheet(ws)

        trades = []
        for trade in rows_data:
            trade_data = {header[i]: trade[i] for i in range(len(trade))}

            trades.append(Trade(trade_data))

        return trades

    def dump_positions(self, positions: list[Position]):
        positions_ws_name = "Positions"

        if positions_ws_name in self.wb.sheetnames:
            self.wb.remove_sheet(self.wb[positions_ws_name])

        self.wb.create_sheet(positions_ws_name, 0)
        ws = self.wb[positions_ws_name]
        ws.append(list(Position.to_row_columns.values()))

        self.wb.active = ws
        for position in positions:
            row = position.to_row()

            ws.append(row)

        self.wb.save(self.path)
